"""
Generate a structured day-by-day DSA Interview Prep plan as an Excel workbook.
Covers all topics needed for MAANG-level coding interviews & LeetCode mastery.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── colour palette ──────────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="2C3E50")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
PHASE_FONT   = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
BODY_FONT    = Font(size=10, name="Calibri")
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="top", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin", color="BDC3C7"),
    right=Side(style="thin", color="BDC3C7"),
    top=Side(style="thin", color="BDC3C7"),
    bottom=Side(style="thin", color="BDC3C7"),
)

DSA_TOPIC_COLORS = {
    "Foundations":                  "1B4F72",
    "Arrays & Hashing":            "117A65",
    "Two Pointers":                "6C3483",
    "Sliding Window":              "B9770E",
    "Stacks & Queues":             "922B21",
    "Binary Search":               "1A5276",
    "Linked Lists":                "0E6655",
    "Sorting":                     "7D3C98",
    "Trees":                       "A04000",
    "Tries":                       "1F618D",
    "Heaps / Priority Queues":     "196F3D",
    "Backtracking":                "7B241C",
    "Graphs":                      "1B4F72",
    "Dynamic Programming":         "6C3483",
    "Greedy":                      "B9770E",
    "Intervals":                   "117A65",
    "Bit Manipulation":            "922B21",
    "Math & Geometry":             "1A5276",
    "Design":                      "196F3D",
    "Mock Interviews & Revision":  "7B241C",
}

# ── The complete DSA curriculum ─────────────────────────────────
# Each entry: (day, category, topic, learning, implementation,
#              leetcode_problems, difficulty, deliverable, hours)
DSA_PLAN: list[tuple] = []
dsa_day = 0

def dsa(cat, topic, learn, impl, problems, diff, deliv, hrs=3.0):
    global dsa_day
    dsa_day += 1
    DSA_PLAN.append((dsa_day, cat, topic, learn, impl, problems, diff, deliv, hrs))

# ── FOUNDATIONS (Days 1-5) ──────────────────────────────────────
F = "Foundations"
dsa(F, "Big-O & Complexity Analysis",
    "Time complexity (O, Ω, Θ), space complexity, amortized analysis, best/avg/worst case",
    "Analyze complexity of 15 code snippets; implement timing benchmarks for O(n) vs O(n²) vs O(n log n)",
    "N/A — theory day",
    "Theory", "complexity_analysis.py with benchmarks + cheat sheet", 3)

dsa(F, "Recursion & Call Stack",
    "Recursion mechanics, base case, call stack, stack overflow, tail recursion",
    "Implement factorial, fibonacci (naive + memo), power(x,n), print all subsets",
    "LC 509 Fibonacci Number (E) • LC 50 Pow(x,n) (M) • LC 779 K-th Symbol in Grammar (M)",
    "Easy-Med", "recursion.py with call stack visualization", 3)

dsa(F, "Hash Tables – Internal Mechanics",
    "Hash functions, collision handling (chaining, open addressing), load factor, resizing",
    "Implement a hash map from scratch with chaining + dynamic resizing",
    "LC 706 Design HashMap (E) • LC 705 Design HashSet (E)",
    "Easy", "hashmap_from_scratch.py with tests", 3)

dsa(F, "Python Data Structures Mastery",
    "list, dict, set, deque, heapq, defaultdict, Counter, OrderedDict — internals & complexity",
    "Benchmark all operations; know exact Big-O of every method",
    "N/A — reference day",
    "Theory", "python_ds_cheatsheet.py with timings", 2)

dsa(F, "Problem-Solving Framework",
    "UMPIRE method (Understand, Match, Plan, Implement, Review, Evaluate), pattern recognition",
    "Solve 5 easy problems using the structured framework; practice thinking aloud",
    "LC 1 Two Sum (E) • LC 242 Valid Anagram (E) • LC 217 Contains Duplicate (E) • LC 26 Remove Duplicates (E) • LC 136 Single Number (E)",
    "Easy", "Framework notes + 5 solutions with written thought process", 3)

# ── ARRAYS & HASHING (Days 6-14) ───────────────────────────────
AH = "Arrays & Hashing"
dsa(AH, "Arrays – Basics & Patterns",
    "Array traversal, prefix sum, suffix sum, in-place modifications",
    "Implement prefix sum array; solve array manipulation problems",
    "LC 238 Product of Array Except Self (M) • LC 303 Range Sum Query (E) • LC 560 Subarray Sum Equals K (M)",
    "Easy-Med", "arrays_basics.py with 3 solutions", 3)

dsa(AH, "Arrays – Frequency Counting",
    "Counter/dict for frequency, bucket sort trick, majority element patterns",
    "Solve frequency-based problems using hash maps",
    "LC 347 Top K Frequent Elements (M) • LC 169 Majority Element (E) • LC 451 Sort Characters By Frequency (M)",
    "Easy-Med", "frequency.py with 3 solutions", 3)

dsa(AH, "Arrays – Grouping & Mapping",
    "Anagram grouping, isomorphic mapping, custom hashing for groups",
    "Solve grouping/mapping problems",
    "LC 49 Group Anagrams (M) • LC 205 Isomorphic Strings (E) • LC 290 Word Pattern (E)",
    "Easy-Med", "grouping.py with 3 solutions", 3)

dsa(AH, "Strings – Core Patterns",
    "String manipulation, palindrome checks, string hashing, character counting",
    "Solve fundamental string problems",
    "LC 125 Valid Palindrome (E) • LC 5 Longest Palindromic Substring (M) • LC 647 Palindromic Substrings (M)",
    "Easy-Med", "strings_core.py with 3 solutions", 3)

dsa(AH, "Strings – Advanced",
    "KMP algorithm, Rabin-Karp, Z-algorithm, longest common subsequence",
    "Implement KMP from scratch; solve string matching problems",
    "LC 28 Find Index of First Occurrence (E) • LC 459 Repeated Substring Pattern (E) • LC 214 Shortest Palindrome (H)",
    "Easy-Hard", "string_matching.py with KMP implementation + 3 solutions", 3)

dsa(AH, "Matrix – Traversal Patterns",
    "Row/col traversal, spiral order, diagonal, transpose, rotation",
    "Solve matrix traversal and manipulation problems",
    "LC 54 Spiral Matrix (M) • LC 48 Rotate Image (M) • LC 73 Set Matrix Zeroes (M)",
    "Medium", "matrix.py with 3 solutions", 3)

dsa(AH, "Hashing – Advanced Problems",
    "Two sum variants, continuous subarray sum, longest consecutive sequence",
    "Solve advanced hash-based problems",
    "LC 128 Longest Consecutive Sequence (M) • LC 523 Continuous Subarray Sum (M) • LC 525 Contiguous Array (M)",
    "Medium", "hashing_advanced.py with 3 solutions", 3)

dsa(AH, "Arrays & Hashing – Hard Problems",
    "Encode/decode, randomized set, first missing positive",
    "Tackle hard-level array/hash problems",
    "LC 41 First Missing Positive (H) • LC 380 Insert Delete GetRandom O(1) (M) • LC 271 Encode and Decode Strings (M)",
    "Med-Hard", "arrays_hard.py with 3 solutions", 3)

dsa(AH, "Arrays & Hashing – Practice Day",
    "Review patterns; solve new problems from scratch",
    "Timed practice: solve 5 problems in 2 hours",
    "LC 36 Valid Sudoku (M) • LC 659 Encode and Decode (M) • LC 14 Longest Common Prefix (E) • LC 118 Pascal's Triangle (E) • LC 287 Find the Duplicate Number (M)",
    "Easy-Med", "5 timed solutions + time log", 3)

# ── TWO POINTERS (Days 15-19) ──────────────────────────────────
TP = "Two Pointers"
dsa(TP, "Two Pointers – Opposite Ends",
    "Sorted array two sum, container with most water, trapping rain water concept",
    "Solve two-pointer problems from both ends",
    "LC 167 Two Sum II (M) • LC 11 Container With Most Water (M) • LC 15 3Sum (M)",
    "Medium", "two_pointers_ends.py with 3 solutions", 3)

dsa(TP, "Two Pointers – Same Direction",
    "Remove duplicates in-place, move zeroes, fast/slow pointer",
    "Solve same-direction two-pointer problems",
    "LC 26 Remove Duplicates from Sorted Array (E) • LC 283 Move Zeroes (E) • LC 27 Remove Element (E) • LC 75 Sort Colors (M)",
    "Easy-Med", "two_pointers_same.py with 4 solutions", 3)

dsa(TP, "Two Pointers – Partition & Dutch National Flag",
    "Three-way partition, Dutch National Flag, segregation problems",
    "Implement 3-way partition; solve segregation problems",
    "LC 75 Sort Colors (M) • LC 86 Partition List (M) • LC 324 Wiggle Sort II (M)",
    "Medium", "partition.py with 3 solutions", 3)

dsa(TP, "Two Pointers – Advanced",
    "4Sum, trapping rain water, multi-pointer techniques",
    "Solve advanced multi-pointer problems",
    "LC 18 4Sum (M) • LC 42 Trapping Rain Water (H) • LC 838 Push Dominoes (M)",
    "Med-Hard", "two_pointers_hard.py with 3 solutions", 3)

dsa(TP, "Two Pointers – Practice Day",
    "Review all two-pointer patterns; timed practice",
    "Timed: solve 4 problems in 90 minutes",
    "LC 680 Valid Palindrome II (E) • LC 881 Boats to Save People (M) • LC 16 3Sum Closest (M) • LC 977 Squares of a Sorted Array (E)",
    "Easy-Med", "4 timed solutions", 3)

# ── SLIDING WINDOW (Days 20-24) ────────────────────────────────
SW = "Sliding Window"
dsa(SW, "Fixed-Size Sliding Window",
    "Window of size K, max sum subarray, averages, fixed window template",
    "Implement fixed-window template; solve classic problems",
    "LC 643 Maximum Average Subarray I (E) • LC 1456 Max Vowels in Substring of Given Length (M) • LC 1343 Number of Sub-arrays of Size K with Avg ≥ Threshold (M)",
    "Easy-Med", "fixed_window.py with template + 3 solutions", 3)

dsa(SW, "Variable-Size Sliding Window",
    "Expand/contract pattern, min window, longest substring without repeats",
    "Implement variable-window template; solve core problems",
    "LC 3 Longest Substring Without Repeating Chars (M) • LC 209 Minimum Size Subarray Sum (M) • LC 1004 Max Consecutive Ones III (M)",
    "Medium", "variable_window.py with template + 3 solutions", 3)

dsa(SW, "Sliding Window with HashMap",
    "Character frequency window, anagram finding, permutation check",
    "Solve hash-augmented sliding window problems",
    "LC 438 Find All Anagrams in a String (M) • LC 567 Permutation in String (M) • LC 30 Substring with Concat of All Words (H)",
    "Med-Hard", "window_hashmap.py with 3 solutions", 3)

dsa(SW, "Sliding Window – Hard Problems",
    "Minimum window substring, sliding window maximum",
    "Solve the hardest sliding window problems",
    "LC 76 Minimum Window Substring (H) • LC 239 Sliding Window Maximum (H) • LC 395 Longest Substring with At Least K Repeating Chars (M)",
    "Hard", "window_hard.py with 3 solutions", 3)

dsa(SW, "Sliding Window – Practice Day",
    "Review templates; timed practice on unseen problems",
    "Timed: solve 4 problems in 90 minutes",
    "LC 424 Longest Repeating Character Replacement (M) • LC 904 Fruit Into Baskets (M) • LC 1208 Get Equal Substrings Within Budget (M) • LC 2024 Maximize the Confusion (M)",
    "Medium", "4 timed solutions", 3)

# ── BINARY SEARCH (Days 25-30) ─────────────────────────────────
BS = "Binary Search"
dsa(BS, "Binary Search – Classic Template",
    "Basic binary search, left/right boundary, search insert position, template I/II/III",
    "Implement 3 binary search templates; know when to use which",
    "LC 704 Binary Search (E) • LC 35 Search Insert Position (E) • LC 374 Guess Number (E)",
    "Easy", "binary_search_templates.py with 3 templates + 3 solutions", 3)

dsa(BS, "Binary Search – On Sorted Arrays",
    "First/last occurrence, search rotated array, find minimum in rotated",
    "Solve sorted-array binary search variants",
    "LC 33 Search in Rotated Sorted Array (M) • LC 153 Find Minimum in Rotated Sorted Array (M) • LC 34 Find First and Last Position (M)",
    "Medium", "bs_sorted.py with 3 solutions", 3)

dsa(BS, "Binary Search – On Answer Space",
    "Binary search on answer (min/max problems), capacity to ship, koko eating bananas",
    "Solve 'binary search on the answer' pattern problems",
    "LC 875 Koko Eating Bananas (M) • LC 1011 Capacity to Ship Packages (M) • LC 410 Split Array Largest Sum (H)",
    "Med-Hard", "bs_on_answer.py with 3 solutions", 3)

dsa(BS, "Binary Search – 2D & Advanced",
    "Search 2D matrix, median of two sorted arrays, split array",
    "Solve 2D and advanced binary search problems",
    "LC 74 Search a 2D Matrix (M) • LC 240 Search a 2D Matrix II (M) • LC 4 Median of Two Sorted Arrays (H)",
    "Med-Hard", "bs_advanced.py with 3 solutions", 3)

dsa(BS, "Binary Search – Practice Day 1",
    "Review templates; apply to new problems",
    "Timed: solve 4 problems in 90 minutes",
    "LC 162 Find Peak Element (M) • LC 540 Single Element in Sorted Array (M) • LC 658 Find K Closest Elements (M) • LC 69 Sqrt(x) (E)",
    "Easy-Med", "4 timed solutions", 3)

dsa(BS, "Binary Search – Practice Day 2",
    "Hard binary search problems; edge cases",
    "Solve hard BS problems with careful edge case handling",
    "LC 154 Find Min in Rotated Sorted Array II (H) • LC 81 Search in Rotated Sorted Array II (M) • LC 378 Kth Smallest Element in Sorted Matrix (M)",
    "Med-Hard", "3 solutions with edge case notes", 3)

# ── STACKS & QUEUES (Days 31-36) ───────────────────────────────
SQ = "Stacks & Queues"
dsa(SQ, "Stack – Fundamentals",
    "Stack operations, LIFO, stack from scratch, parentheses matching, evaluation",
    "Implement stack from scratch; solve classic stack problems",
    "LC 20 Valid Parentheses (E) • LC 155 Min Stack (M) • LC 150 Evaluate Reverse Polish Notation (M)",
    "Easy-Med", "stack_basics.py with stack impl + 3 solutions", 3)

dsa(SQ, "Monotonic Stack",
    "Next greater/smaller element, monotonic increasing/decreasing stack template",
    "Implement monotonic stack template; solve core problems",
    "LC 496 Next Greater Element I (E) • LC 739 Daily Temperatures (M) • LC 503 Next Greater Element II (M)",
    "Easy-Med", "monotonic_stack.py with template + 3 solutions", 3)

dsa(SQ, "Monotonic Stack – Advanced",
    "Largest rectangle in histogram, maximal rectangle, stock span",
    "Solve hard monotonic stack problems",
    "LC 84 Largest Rectangle in Histogram (H) • LC 85 Maximal Rectangle (H) • LC 901 Online Stock Span (M)",
    "Med-Hard", "monotonic_stack_hard.py with 3 solutions", 3)

dsa(SQ, "Queue & Deque",
    "Queue from scratch, circular queue, deque operations, BFS primer",
    "Implement circular queue; solve queue-based problems",
    "LC 622 Design Circular Queue (M) • LC 933 Number of Recent Calls (E) • LC 346 Moving Average (E)",
    "Easy-Med", "queue.py with circular queue impl + 3 solutions", 3)

dsa(SQ, "Stack – Application Problems",
    "Decode string, asteroid collision, simplify path, calculator",
    "Solve application-heavy stack problems",
    "LC 394 Decode String (M) • LC 735 Asteroid Collision (M) • LC 71 Simplify Path (M)",
    "Medium", "stack_apps.py with 3 solutions", 3)

dsa(SQ, "Stack & Queue – Practice Day",
    "Review all stack/queue patterns; timed practice",
    "Timed: solve 4 problems in 100 minutes",
    "LC 224 Basic Calculator (H) • LC 232 Implement Queue using Stacks (E) • LC 225 Implement Stack using Queues (E) • LC 402 Remove K Digits (M)",
    "Easy-Hard", "4 timed solutions", 3)

# ── LINKED LISTS (Days 37-42) ──────────────────────────────────
LL = "Linked Lists"
dsa(LL, "Singly Linked List – From Scratch",
    "Node class, insert, delete, search, reverse, LL vs array trade-offs",
    "Implement singly linked list class with all operations",
    "LC 206 Reverse Linked List (E) • LC 21 Merge Two Sorted Lists (E) • LC 707 Design Linked List (M)",
    "Easy-Med", "singly_ll.py with full implementation + 3 solutions", 3)

dsa(LL, "Linked List – Two Pointer Patterns",
    "Fast/slow pointer (Floyd's), cycle detection, middle element, nth from end",
    "Implement cycle detection and find-middle; solve problems",
    "LC 141 Linked List Cycle (E) • LC 142 Linked List Cycle II (M) • LC 876 Middle of the Linked List (E) • LC 19 Remove Nth Node From End (M)",
    "Easy-Med", "ll_two_pointers.py with 4 solutions", 3)

dsa(LL, "Linked List – Reversal Patterns",
    "Reverse sublist, reverse in K-groups, swap pairs, palindrome LL",
    "Solve reversal-based linked list problems",
    "LC 92 Reverse Linked List II (M) • LC 25 Reverse Nodes in k-Group (H) • LC 234 Palindrome Linked List (E)",
    "Easy-Hard", "ll_reversal.py with 3 solutions", 3)

dsa(LL, "Linked List – Merge & Sort",
    "Merge K sorted lists, sort linked list (merge sort), intersection",
    "Implement merge sort for LL; solve merge problems",
    "LC 23 Merge k Sorted Lists (H) • LC 148 Sort List (M) • LC 160 Intersection of Two LLs (E)",
    "Easy-Hard", "ll_merge.py with 3 solutions", 3)

dsa(LL, "Doubly Linked List & LRU Cache",
    "Doubly LL, LRU cache with DLL + HashMap, LFU concept",
    "Implement LRU Cache from scratch (DLL + HashMap)",
    "LC 146 LRU Cache (M) • LC 460 LFU Cache (H) • LC 138 Copy List with Random Pointer (M)",
    "Med-Hard", "lru_cache.py with from-scratch implementation + 3 solutions", 4)

dsa(LL, "Linked Lists – Practice Day",
    "Review all LL patterns; timed practice",
    "Timed: solve 4 problems in 90 minutes",
    "LC 2 Add Two Numbers (M) • LC 143 Reorder List (M) • LC 287 Find Duplicate Number (M) • LC 61 Rotate List (M)",
    "Medium", "4 timed solutions", 3)

# ── SORTING (Days 43-46) ───────────────────────────────────────
SO = "Sorting"
dsa(SO, "Sorting – Comparison-Based (Part 1)",
    "Bubble, selection, insertion sort — mechanics, stability, Time/Space",
    "Implement all 3 sorts from scratch; benchmark on various inputs",
    "N/A — implementation day",
    "Theory", "basic_sorts.py with implementations + benchmarks", 3)

dsa(SO, "Sorting – Efficient Sorts",
    "Merge sort, quick sort (Lomuto + Hoare), heap sort — divide & conquer",
    "Implement merge sort, quick sort (both partitions), heap sort from scratch",
    "LC 912 Sort an Array (M) • LC 215 Kth Largest Element (M)",
    "Medium", "efficient_sorts.py with 3 implementations + 2 solutions", 4)

dsa(SO, "Sorting – Non-Comparison & Special",
    "Counting sort, radix sort, bucket sort, Tim sort concepts; when to use which",
    "Implement counting + radix sort from scratch",
    "LC 274 H-Index (M) • LC 164 Maximum Gap (M) • LC 179 Largest Number (M)",
    "Medium", "non_comparison_sorts.py with implementations + 3 solutions", 3)

dsa(SO, "Sorting – Application Problems",
    "Custom sort keys, interval scheduling, merge intervals, meeting rooms",
    "Solve sorting-based application problems",
    "LC 56 Merge Intervals (M) • LC 252 Meeting Rooms (E) • LC 253 Meeting Rooms II (M) • LC 973 K Closest Points to Origin (M)",
    "Easy-Med", "sort_apps.py with 4 solutions", 3)

# ── TREES (Days 47-58) ─────────────────────────────────────────
TR = "Trees"
dsa(TR, "Binary Tree – Traversals",
    "Inorder, preorder, postorder (recursive + iterative), level-order (BFS)",
    "Implement all 4 traversals both recursively and iteratively",
    "LC 94 Binary Tree Inorder Traversal (E) • LC 144 Preorder (E) • LC 145 Postorder (E) • LC 102 Level Order (M)",
    "Easy-Med", "tree_traversals.py with 6 implementations + 4 solutions", 3)

dsa(TR, "Binary Tree – Basic Properties",
    "Height, depth, diameter, balanced check, symmetric, invert",
    "Solve fundamental binary tree property problems",
    "LC 104 Maximum Depth (E) • LC 543 Diameter of Binary Tree (E) • LC 110 Balanced Binary Tree (E) • LC 226 Invert Binary Tree (E)",
    "Easy", "tree_properties.py with 4 solutions", 3)

dsa(TR, "Binary Tree – Path Problems",
    "Root-to-leaf paths, path sum, max path sum, LCA",
    "Solve path-based tree problems",
    "LC 112 Path Sum (E) • LC 113 Path Sum II (M) • LC 124 Binary Tree Max Path Sum (H) • LC 236 LCA of Binary Tree (M)",
    "Easy-Hard", "tree_paths.py with 4 solutions", 3)

dsa(TR, "Binary Tree – Construction & Serialization",
    "Build tree from traversals, serialize/deserialize, codec",
    "Solve tree construction and serialization problems",
    "LC 105 Construct BT from Preorder and Inorder (M) • LC 106 from Inorder and Postorder (M) • LC 297 Serialize and Deserialize BT (H)",
    "Med-Hard", "tree_construct.py with 3 solutions", 3)

dsa(TR, "Binary Tree – Advanced",
    "Vertical order, right side view, zigzag, boundary traversal",
    "Solve advanced traversal-variant problems",
    "LC 199 Binary Tree Right Side View (M) • LC 103 Zigzag Level Order (M) • LC 314 Binary Tree Vertical Order (M) • LC 987 Vertical Order Traversal (H)",
    "Med-Hard", "tree_advanced.py with 4 solutions", 3)

dsa(TR, "Binary Search Tree – Fundamentals",
    "BST property, search, insert, delete, in-order is sorted, validate BST",
    "Implement BST with insert/delete/search from scratch",
    "LC 700 Search in a BST (E) • LC 701 Insert into BST (M) • LC 450 Delete Node in BST (M) • LC 98 Validate BST (M)",
    "Easy-Med", "bst.py with full implementation + 4 solutions", 3)

dsa(TR, "BST – Advanced Problems",
    "Kth smallest, BST iterator, convert sorted array to BST, recover BST",
    "Solve advanced BST problems using in-order properties",
    "LC 230 Kth Smallest Element in BST (M) • LC 173 BST Iterator (M) • LC 108 Convert Sorted Array to BST (E) • LC 99 Recover BST (M)",
    "Easy-Med", "bst_advanced.py with 4 solutions", 3)

dsa(TR, "N-ary Trees & Special Trees",
    "N-ary tree traversal, trie preview, segment tree overview, Fenwick tree concept",
    "Implement N-ary tree traversal and basic segment tree",
    "LC 589 N-ary Tree Preorder (E) • LC 590 N-ary Tree Postorder (E) • LC 429 N-ary Tree Level Order (M)",
    "Easy-Med", "nary_trees.py with 3 solutions + segment tree skeleton", 3)

dsa(TR, "Trees – Practice Day 1",
    "Review all tree patterns; timed practice",
    "Timed: solve 5 problems in 2 hours",
    "LC 572 Subtree of Another Tree (E) • LC 235 LCA of BST (M) • LC 1448 Count Good Nodes (M) • LC 437 Path Sum III (M) • LC 662 Maximum Width of BT (M)",
    "Easy-Med", "5 timed solutions", 3)

dsa(TR, "Trees – Practice Day 2",
    "Hard tree problems; complex recursion",
    "Solve challenging tree problems",
    "LC 968 Binary Tree Cameras (H) • LC 834 Sum of Distances in Tree (H) • LC 979 Distribute Coins in BT (M) • LC 863 All Nodes Distance K (M)",
    "Med-Hard", "4 solutions with detailed analysis", 3)

# ── TRIES (Days 59-61) ─────────────────────────────────────────
TI = "Tries"
dsa(TI, "Trie – Implementation & Basics",
    "Trie node structure, insert, search, startsWith, space-time analysis",
    "Implement Trie from scratch; understand memory trade-offs",
    "LC 208 Implement Trie (M) • LC 211 Design Add and Search Words (M)",
    "Medium", "trie.py with from-scratch implementation + 2 solutions", 3)

dsa(TI, "Trie – Word Search & Advanced",
    "Word search in board with trie, auto-complete, longest word",
    "Solve trie-augmented search problems",
    "LC 212 Word Search II (H) • LC 648 Replace Words (M) • LC 720 Longest Word in Dictionary (M)",
    "Med-Hard", "trie_advanced.py with 3 solutions", 3)

dsa(TI, "Trie – Practice Day",
    "Trie pattern review; prefix-based problems",
    "Solve additional trie and prefix problems",
    "LC 14 Longest Common Prefix (E) • LC 1268 Search Suggestions System (M) • LC 677 Map Sum Pairs (M)",
    "Easy-Med", "3 solutions", 3)

# ── HEAPS / PRIORITY QUEUES (Days 62-67) ───────────────────────
HP = "Heaps / Priority Queues"
dsa(HP, "Heap – From Scratch",
    "Min-heap, max-heap, heapify, push, pop, sift up/down, build heap O(n)",
    "Implement min-heap and max-heap from scratch",
    "LC 703 Kth Largest Element in Stream (E) • N/A",
    "Easy", "heap_from_scratch.py with full implementation + 1 solution", 3)

dsa(HP, "Heap – Top K Pattern",
    "Top K elements, K closest, K most frequent using heap",
    "Solve top-K problems using heaps",
    "LC 215 Kth Largest Element in Array (M) • LC 347 Top K Frequent Elements (M) • LC 973 K Closest Points to Origin (M)",
    "Medium", "heap_topk.py with 3 solutions", 3)

dsa(HP, "Heap – Two Heaps Pattern",
    "Median maintenance with two heaps (max-heap + min-heap), sliding window median",
    "Implement find-median-from-stream",
    "LC 295 Find Median from Data Stream (H) • LC 480 Sliding Window Median (H)",
    "Hard", "two_heaps.py with 2 solutions", 3)

dsa(HP, "Heap – Merge Pattern",
    "Merge k sorted lists/arrays using heap, smallest range",
    "Solve merge-based heap problems",
    "LC 23 Merge k Sorted Lists (H) • LC 378 Kth Smallest Element in Sorted Matrix (M) • LC 632 Smallest Range Covering Elements (H)",
    "Med-Hard", "heap_merge.py with 3 solutions", 3)

dsa(HP, "Heap – Scheduling & Greedy",
    "Task scheduler, reorganize string, meeting rooms with heap",
    "Solve scheduling problems using priority queues",
    "LC 621 Task Scheduler (M) • LC 767 Reorganize String (M) • LC 1834 Single-Threaded CPU (M)",
    "Medium", "heap_scheduling.py with 3 solutions", 3)

dsa(HP, "Heaps – Practice Day",
    "Review heap patterns; timed practice",
    "Timed: solve 4 problems in 100 minutes",
    "LC 355 Design Twitter (M) • LC 659 Split Array into Consecutive Subsequences (M) • LC 1046 Last Stone Weight (E) • LC 502 IPO (H)",
    "Easy-Hard", "4 timed solutions", 3)

# ── BACKTRACKING (Days 68-73) ──────────────────────────────────
BT = "Backtracking"
dsa(BT, "Backtracking – Template & Subsets",
    "Backtracking template, state space tree, subsets, subsets with duplicates",
    "Implement backtracking template; solve subset problems",
    "LC 78 Subsets (M) • LC 90 Subsets II (M) • LC 784 Letter Case Permutation (M)",
    "Medium", "backtrack_subsets.py with template + 3 solutions", 3)

dsa(BT, "Backtracking – Permutations",
    "Permutations, permutations with duplicates, swap-based approach",
    "Solve permutation problems",
    "LC 46 Permutations (M) • LC 47 Permutations II (M) • LC 31 Next Permutation (M)",
    "Medium", "backtrack_perms.py with 3 solutions", 3)

dsa(BT, "Backtracking – Combinations",
    "Combination sum variants, prune early, deduplication strategies",
    "Solve combination problems with pruning",
    "LC 39 Combination Sum (M) • LC 40 Combination Sum II (M) • LC 216 Combination Sum III (M) • LC 77 Combinations (M)",
    "Medium", "backtrack_combos.py with 4 solutions", 3)

dsa(BT, "Backtracking – Grid & Board",
    "N-Queens, Sudoku solver, word search, island variants",
    "Solve grid-based backtracking problems",
    "LC 51 N-Queens (H) • LC 37 Sudoku Solver (H) • LC 79 Word Search (M)",
    "Med-Hard", "backtrack_grid.py with 3 solutions", 3)

dsa(BT, "Backtracking – String & Partition",
    "Palindrome partitioning, generate parentheses, letter combinations phone",
    "Solve string partitioning and generation problems",
    "LC 131 Palindrome Partitioning (M) • LC 22 Generate Parentheses (M) • LC 17 Letter Combinations of Phone Number (M) • LC 93 Restore IP Addresses (M)",
    "Medium", "backtrack_strings.py with 4 solutions", 3)

dsa(BT, "Backtracking – Practice Day",
    "Review backtracking template; timed practice on hard problems",
    "Timed: solve 4 problems in 2 hours",
    "LC 52 N-Queens II (H) • LC 698 Partition to K Equal Sum Subsets (M) • LC 473 Matchsticks to Square (M) • LC 1849 Splitting a String into Descending Consecutive Values (M)",
    "Med-Hard", "4 timed solutions", 3)

# ── GRAPHS (Days 74-87) ───────────────────────────────────────
GR = "Graphs"
dsa(GR, "Graph – Representations & Basics",
    "Adjacency list, adjacency matrix, edge list, directed/undirected, weighted, graph terminology",
    "Implement graph class with adj list + adj matrix; convert between representations",
    "N/A — implementation day",
    "Theory", "graph_repr.py with Graph class and conversion methods", 3)

dsa(GR, "Graph – BFS",
    "BFS traversal, level-by-level, shortest path in unweighted graph, BFS template",
    "Implement BFS from scratch; solve BFS problems",
    "LC 733 Flood Fill (E) • LC 994 Rotting Oranges (M) • LC 542 01 Matrix (M)",
    "Easy-Med", "graph_bfs.py with BFS template + 3 solutions", 3)

dsa(GR, "Graph – DFS",
    "DFS traversal (recursive + iterative), connected components, visited set",
    "Implement DFS from scratch; solve DFS problems",
    "LC 200 Number of Islands (M) • LC 695 Max Area of Island (M) • LC 547 Number of Provinces (M)",
    "Medium", "graph_dfs.py with DFS template + 3 solutions", 3)

dsa(GR, "Graph – Grid as Graph",
    "4-directional and 8-directional traversal, boundary conditions, multi-source BFS",
    "Solve grid-based graph problems",
    "LC 1091 Shortest Path in Binary Matrix (M) • LC 286 Walls and Gates (M) • LC 417 Pacific Atlantic Water Flow (M)",
    "Medium", "grid_graph.py with 3 solutions", 3)

dsa(GR, "Graph – Cycle Detection",
    "Cycle in undirected (DFS, Union-Find), cycle in directed (coloring, topological sort)",
    "Implement cycle detection for both graph types",
    "LC 207 Course Schedule (M) • LC 210 Course Schedule II (M) • LC 802 Find Eventual Safe States (M)",
    "Medium", "cycle_detection.py with 3 solutions", 3)

dsa(GR, "Topological Sort",
    "Kahn's algorithm (BFS), DFS-based topo sort, DAG, dependency resolution",
    "Implement both topo sort approaches from scratch",
    "LC 210 Course Schedule II (M) • LC 269 Alien Dictionary (H) • LC 310 Minimum Height Trees (M)",
    "Med-Hard", "topological_sort.py with 2 implementations + 3 solutions", 3)

dsa(GR, "Graph – Union-Find / Disjoint Set",
    "Union by rank, path compression, connected components, Kruskal's integration",
    "Implement Union-Find from scratch with optimizations",
    "LC 323 Number of Connected Components (M) • LC 684 Redundant Connection (M) • LC 721 Accounts Merge (M)",
    "Medium", "union_find.py with from-scratch impl + 3 solutions", 3)

dsa(GR, "Graph – Shortest Path (Dijkstra)",
    "Dijkstra's algorithm, priority queue, weighted graph, non-negative weights",
    "Implement Dijkstra from scratch; solve shortest path problems",
    "LC 743 Network Delay Time (M) • LC 1514 Path with Maximum Probability (M) • LC 787 Cheapest Flights Within K Stops (M)",
    "Medium", "dijkstra.py with from-scratch impl + 3 solutions", 3)

dsa(GR, "Graph – Bellman-Ford & Floyd-Warshall",
    "Bellman-Ford (negative weights), Floyd-Warshall (all pairs), negative cycle detection",
    "Implement both algorithms from scratch",
    "LC 787 Cheapest Flights Within K Stops (M) • LC 1334 Find the City (M)",
    "Medium", "bellman_floyd.py with 2 implementations + 2 solutions", 3)

dsa(GR, "Graph – MST (Prim & Kruskal)",
    "Minimum Spanning Tree, Prim's (greedy + heap), Kruskal's (sort + UF)",
    "Implement both MST algorithms from scratch",
    "LC 1584 Min Cost to Connect All Points (M) • LC 1135 Connecting Cities With Minimum Cost (M)",
    "Medium", "mst.py with 2 implementations + 2 solutions", 3)

dsa(GR, "Graph – Bipartite & Coloring",
    "Bipartite check (BFS/DFS coloring), graph coloring, odd cycle",
    "Implement bipartite checking; solve coloring problems",
    "LC 785 Is Graph Bipartite? (M) • LC 886 Possible Bipartition (M) • LC 1042 Flower Planting With No Adjacent (M)",
    "Medium", "bipartite.py with 3 solutions", 3)

dsa(GR, "Graph – Hard Problems",
    "Word ladder, clone graph, alien dictionary, critical connections (bridges)",
    "Solve hard graph problems; implement Tarjan's bridge-finding",
    "LC 127 Word Ladder (H) • LC 133 Clone Graph (M) • LC 1192 Critical Connections (H)",
    "Med-Hard", "graph_hard.py with 3 solutions", 3)

dsa(GR, "Graph – Practice Day 1",
    "Review BFS/DFS/Union-Find patterns; timed practice",
    "Timed: solve 5 problems in 2.5 hours",
    "LC 1319 Number of Ops to Make Network Connected (M) • LC 841 Keys and Rooms (M) • LC 1162 As Far From Land as Possible (M) • LC 909 Snakes and Ladders (M) • LC 1926 Nearest Exit from Entrance (M)",
    "Medium", "5 timed solutions", 3)

dsa(GR, "Graph – Practice Day 2",
    "Advanced graph patterns; hard problems",
    "Solve hard graph problems",
    "LC 332 Reconstruct Itinerary (H) • LC 1168 Optimize Water Distribution (H) • LC 778 Swim in Rising Water (H) • LC 1631 Path With Minimum Effort (M)",
    "Med-Hard", "4 solutions with detailed analysis", 3)

# ── DYNAMIC PROGRAMMING (Days 88-107) ─────────────────────────
DP = "Dynamic Programming"
dsa(DP, "DP – Theory & 1D Basics",
    "Overlapping subproblems, optimal substructure, memoization vs tabulation, state definition",
    "Solve basic 1D DP problems with both approaches",
    "LC 70 Climbing Stairs (E) • LC 198 House Robber (M) • LC 509 Fibonacci Number (E) • LC 746 Min Cost Climbing Stairs (E)",
    "Easy-Med", "dp_1d_basics.py with memo + tabulation for each + 4 solutions", 3)

dsa(DP, "DP – 1D Classic Problems",
    "Coin change, word break, decode ways, longest increasing subsequence",
    "Solve classic 1D DP problems",
    "LC 322 Coin Change (M) • LC 139 Word Break (M) • LC 91 Decode Ways (M) • LC 300 Longest Increasing Subsequence (M)",
    "Medium", "dp_1d_classic.py with 4 solutions", 3)

dsa(DP, "DP – House Robber & Jump Game Variants",
    "House robber variants, jump game variants, state transitions",
    "Solve DP problems with simple state transitions",
    "LC 198 House Robber (M) • LC 213 House Robber II (M) • LC 55 Jump Game (M) • LC 45 Jump Game II (M)",
    "Medium", "dp_robber_jump.py with 4 solutions", 3)

dsa(DP, "DP – 2D Grid Problems",
    "Unique paths, minimum path sum, grid traversal DP",
    "Solve 2D grid DP problems",
    "LC 62 Unique Paths (M) • LC 63 Unique Paths II (M) • LC 64 Minimum Path Sum (M) • LC 120 Triangle (M)",
    "Medium", "dp_grid.py with 4 solutions", 3)

dsa(DP, "DP – String Problems (Part 1)",
    "Longest common subsequence, edit distance, state: dp[i][j] = first i chars of s1, first j of s2",
    "Solve string DP problems",
    "LC 1143 Longest Common Subsequence (M) • LC 72 Edit Distance (M) • LC 583 Delete Operation for Two Strings (M)",
    "Medium", "dp_strings1.py with 3 solutions", 3)

dsa(DP, "DP – String Problems (Part 2)",
    "Longest palindromic subsequence, distinct subsequences, interleaving strings",
    "Solve advanced string DP problems",
    "LC 516 Longest Palindromic Subsequence (M) • LC 115 Distinct Subsequences (H) • LC 97 Interleaving String (M)",
    "Med-Hard", "dp_strings2.py with 3 solutions", 3)

dsa(DP, "DP – Knapsack Pattern",
    "0/1 knapsack, unbounded knapsack, subset sum, partition equal subset",
    "Implement 0/1 and unbounded knapsack from scratch; solve variants",
    "LC 416 Partition Equal Subset Sum (M) • LC 494 Target Sum (M) • LC 518 Coin Change II (M) • LC 474 Ones and Zeroes (M)",
    "Medium", "dp_knapsack.py with knapsack impl + 4 solutions", 3)

dsa(DP, "DP – Decision Making",
    "Buy/sell stock variants, decision DP with states, state machine approach",
    "Solve stock trading DP problems",
    "LC 121 Best Time Buy/Sell Stock (E) • LC 122 Stock II (M) • LC 123 Stock III (H) • LC 188 Stock IV (H) • LC 309 Stock with Cooldown (M)",
    "Easy-Hard", "dp_stocks.py with 5 solutions using state machine", 3)

dsa(DP, "DP – Interval DP",
    "Burst balloons, matrix chain multiplication, stone game, min cost to merge",
    "Solve interval DP problems",
    "LC 312 Burst Balloons (H) • LC 1039 Minimum Score Triangulation (M) • LC 877 Stone Game (M)",
    "Med-Hard", "dp_interval.py with 3 solutions", 3)

dsa(DP, "DP – Bitmask DP",
    "Bitmask as state, subsets enumeration, Hamiltonian path, assignment problem",
    "Solve bitmask DP problems",
    "LC 526 Beautiful Arrangement (M) • LC 698 Partition to K Equal Sum Subsets (M) • LC 1125 Smallest Sufficient Team (H)",
    "Med-Hard", "dp_bitmask.py with 3 solutions", 3)

dsa(DP, "DP – Tree DP",
    "DP on trees, re-rooting technique, house robber III, binary tree cameras",
    "Solve tree DP problems",
    "LC 337 House Robber III (M) • LC 968 Binary Tree Cameras (H) • LC 834 Sum of Distances in Tree (H)",
    "Med-Hard", "dp_tree.py with 3 solutions", 3)

dsa(DP, "DP – on Digits",
    "Digit DP, count numbers in range with property, leading zeros",
    "Solve digit DP problems",
    "LC 233 Number of Digit One (H) • LC 902 Numbers At Most N Given Digit Set (H) • LC 357 Count Numbers with Unique Digits (M)",
    "Med-Hard", "dp_digits.py with 3 solutions", 3)

dsa(DP, "DP – Optimization (Space)",
    "Space optimization: rolling array, 2-row trick, O(n) from O(n²)",
    "Revisit 5 solved DP problems and optimize space",
    "Optimize: Unique Paths, LCS, Edit Distance, Knapsack, Coin Change",
    "Medium", "dp_space_opt.py with 5 optimized solutions", 3)

dsa(DP, "DP – Mixed Practice Day 1",
    "Review all DP patterns; timed practice",
    "Timed: solve 5 problems in 2.5 hours",
    "LC 264 Ugly Number II (M) • LC 279 Perfect Squares (M) • LC 343 Integer Break (M) • LC 152 Maximum Product Subarray (M) • LC 377 Combination Sum IV (M)",
    "Medium", "5 timed solutions", 3)

dsa(DP, "DP – Mixed Practice Day 2",
    "Hard DP problems; complex state transitions",
    "Solve hard DP problems",
    "LC 10 Regular Expression Matching (H) • LC 44 Wildcard Matching (H) • LC 87 Scramble String (H)",
    "Hard", "3 hard solutions with detailed state transition analysis", 3)

dsa(DP, "DP – Mixed Practice Day 3",
    "Contest-level DP problems; DP with other techniques",
    "Solve hard DP combinations",
    "LC 1547 Min Cost to Cut a Stick (H) • LC 1335 Min Difficulty of Job Schedule (H) • LC 1478 Allocate Mailboxes (H)",
    "Hard", "3 hard solutions with approach comparison", 3)

dsa(DP, "DP – Comprehensive Review",
    "Review ALL DP patterns: 1D, 2D, string, knapsack, interval, tree, bitmask, digit",
    "Create DP pattern cheat sheet; solve 3 unseen problems",
    "3 new medium/hard problems (pick from blind 75 list)",
    "Med-Hard", "DP pattern cheat sheet + 3 new solutions", 3)

# ── GREEDY (Days 108-113) ──────────────────────────────────────
GD = "Greedy"
dsa(GD, "Greedy – Theory & Easy Problems",
    "Greedy choice property, proof strategies (exchange argument, stays ahead), greedy vs DP",
    "Solve easy greedy problems; practice proving correctness",
    "LC 455 Assign Cookies (E) • LC 860 Lemonade Change (E) • LC 1005 Maximize Sum After K Negations (E)",
    "Easy", "greedy_basics.py with proofs + 3 solutions", 3)

dsa(GD, "Greedy – Interval Problems",
    "Activity selection, interval scheduling, non-overlapping intervals",
    "Solve greedy interval problems",
    "LC 435 Non-overlapping Intervals (M) • LC 452 Minimum Arrows to Burst Balloons (M) • LC 56 Merge Intervals (M)",
    "Medium", "greedy_intervals.py with 3 solutions", 3)

dsa(GD, "Greedy – Scheduling & Arrangement",
    "Job sequencing, task scheduler, gas station, candy distribution",
    "Solve greedy scheduling and arrangement problems",
    "LC 134 Gas Station (M) • LC 135 Candy (H) • LC 621 Task Scheduler (M) • LC 406 Queue Reconstruction by Height (M)",
    "Med-Hard", "greedy_scheduling.py with 4 solutions", 3)

dsa(GD, "Greedy – String Problems",
    "Partition labels, remove K digits, smallest subsequence, rearrange",
    "Solve greedy string problems",
    "LC 763 Partition Labels (M) • LC 402 Remove K Digits (M) • LC 316 Remove Duplicate Letters (M) • LC 678 Valid Parenthesis String (M)",
    "Medium", "greedy_strings.py with 4 solutions", 3)

dsa(GD, "Greedy – Advanced & Math",
    "Jump game, hand of straights, maximum subarray (Kadane's), redistribute",
    "Solve advanced greedy problems",
    "LC 55 Jump Game (M) • LC 846 Hand of Straights (M) • LC 53 Maximum Subarray (M) • LC 1899 Merge Triplets (M)",
    "Medium", "greedy_advanced.py with 4 solutions", 3)

dsa(GD, "Greedy – Practice Day",
    "Review greedy patterns; timed practice",
    "Timed: solve 4 problems in 100 minutes",
    "LC 948 Bag of Tokens (M) • LC 1353 Maximum Number of Events (M) • LC 659 Split Array into Consecutive Subseq (M) • LC 2131 Longest Palindrome by Concat (M)",
    "Medium", "4 timed solutions", 3)

# ── INTERVALS (Days 114-116) ───────────────────────────────────
IV = "Intervals"
dsa(IV, "Intervals – Core Patterns",
    "Merge, insert, overlapping check, sweep line basics",
    "Solve core interval problems",
    "LC 56 Merge Intervals (M) • LC 57 Insert Interval (M) • LC 986 Interval List Intersections (M)",
    "Medium", "intervals_core.py with 3 solutions", 3)

dsa(IV, "Intervals – Meeting Rooms & Sweep Line",
    "Meeting rooms min, calendar booking, sweep line algorithm",
    "Solve meeting-room and sweep-line problems",
    "LC 253 Meeting Rooms II (M) • LC 729 My Calendar I (M) • LC 1288 Remove Covered Intervals (M)",
    "Medium", "intervals_sweep.py with 3 solutions", 3)

dsa(IV, "Intervals – Practice Day",
    "Review interval patterns; timed practice",
    "Timed: solve 3 problems in 75 minutes",
    "LC 352 Data Stream as Disjoint Intervals (H) • LC 759 Employee Free Time (H) • LC 1235 Maximum Profit in Job Scheduling (H)",
    "Hard", "3 timed solutions", 3)

# ── BIT MANIPULATION (Days 117-120) ────────────────────────────
BM = "Bit Manipulation"
dsa(BM, "Bit Manipulation – Basics",
    "AND, OR, XOR, NOT, left/right shift, bit tricks (check/set/clear/toggle bit), two's complement",
    "Implement bit manipulation utility functions; solve basic problems",
    "LC 136 Single Number (E) • LC 191 Number of 1 Bits (E) • LC 338 Counting Bits (E) • LC 231 Power of Two (E)",
    "Easy", "bit_basics.py with utility functions + 4 solutions", 3)

dsa(BM, "Bit Manipulation – Intermediate",
    "XOR patterns (missing/duplicate), bit masking, subset enumeration",
    "Solve XOR-based and bitmask problems",
    "LC 137 Single Number II (M) • LC 260 Single Number III (M) • LC 371 Sum of Two Integers (M) • LC 268 Missing Number (E)",
    "Easy-Med", "bit_intermediate.py with 4 solutions", 3)

dsa(BM, "Bit Manipulation – Advanced",
    "Bit DP overlap, binary representation tricks, max XOR",
    "Solve advanced bit manipulation problems",
    "LC 421 Maximum XOR of Two Numbers in Array (M) • LC 201 Bitwise AND of Number Range (M) • LC 318 Maximum Product of Word Lengths (M)",
    "Medium", "bit_advanced.py with 3 solutions", 3)

dsa(BM, "Bit Manipulation – Practice Day",
    "Review bit patterns; timed practice",
    "Timed: solve 4 problems in 90 minutes",
    "LC 190 Reverse Bits (E) • LC 461 Hamming Distance (E) • LC 29 Divide Two Integers (M) • LC 89 Gray Code (M)",
    "Easy-Med", "4 timed solutions", 3)

# ── MATH & GEOMETRY (Days 121-124) ─────────────────────────────
MG = "Math & Geometry"
dsa(MG, "Math – Number Theory",
    "GCD (Euclidean), LCM, primes (sieve), modular arithmetic, fast exponentiation",
    "Implement sieve of Eratosthenes and modular exponentiation",
    "LC 204 Count Primes (M) • LC 1492 The kth Factor of n (M) • LC 50 Pow(x,n) (M)",
    "Medium", "math_number_theory.py with implementations + 3 solutions", 3)

dsa(MG, "Math – Combinatorics & Probability",
    "Combinations, permutations, Pascal's triangle, expected value, reservoir sampling",
    "Implement combinations and reservoir sampling",
    "LC 118 Pascal's Triangle (E) • LC 62 Unique Paths (M) • LC 382 Linked List Random Node (M)",
    "Easy-Med", "math_combinatorics.py with 3 solutions", 3)

dsa(MG, "Geometry – Points & Lines",
    "Distance, slope, collinear check, convex hull concept, rectangle overlap",
    "Solve geometry problems",
    "LC 149 Max Points on a Line (H) • LC 223 Rectangle Area (M) • LC 836 Rectangle Overlap (E)",
    "Easy-Hard", "geometry.py with 3 solutions", 3)

dsa(MG, "Math & Geometry – Practice",
    "Matrix rotation, spiral, robot problems, math tricks",
    "Solve math/geometry practice problems",
    "LC 43 Multiply Strings (M) • LC 7 Reverse Integer (M) • LC 9 Palindrome Number (E) • LC 66 Plus One (E)",
    "Easy-Med", "4 timed solutions", 3)

# ── DESIGN (Days 125-130) ──────────────────────────────────────
DS = "Design"
dsa(DS, "Design – Stack & Queue Based",
    "MinStack, MaxStack, stack with increment, circular deque",
    "Implement design problems using stacks/queues",
    "LC 155 Min Stack (M) • LC 716 Max Stack (H) • LC 1381 Design Stack With Increment Op (M)",
    "Med-Hard", "design_stack.py with 3 solutions", 3)

dsa(DS, "Design – HashMap & Set Based",
    "RandomizedSet, time-based KV store, snapshot array",
    "Implement advanced hash-based data structures",
    "LC 380 Insert Delete GetRandom O(1) (M) • LC 981 Time Based Key-Value Store (M) • LC 1146 Snapshot Array (M)",
    "Medium", "design_hash.py with 3 solutions", 3)

dsa(DS, "Design – Linked List Based",
    "LRU Cache, LFU Cache, skip list concepts",
    "Implement LRU and LFU from scratch",
    "LC 146 LRU Cache (M) • LC 460 LFU Cache (H)",
    "Med-Hard", "design_ll.py with 2 from-scratch implementations", 4)

dsa(DS, "Design – Iterator & Stream",
    "Flatten nested list, peeking iterator, iterator patterns",
    "Implement iterator-based designs",
    "LC 341 Flatten Nested List Iterator (M) • LC 284 Peeking Iterator (M) • LC 251 Flatten 2D Vector (M)",
    "Medium", "design_iterator.py with 3 solutions", 3)

dsa(DS, "Design – Complex Systems",
    "Twitter feed, in-memory file system, browser history",
    "Implement complex OO design problems",
    "LC 355 Design Twitter (M) • LC 588 Design In-Memory File System (H) • LC 1472 Design Browser History (M)",
    "Med-Hard", "design_systems.py with 3 solutions", 3)

dsa(DS, "Design – Practice Day",
    "Review design patterns; timed practice",
    "Timed: solve 3 design problems in 2 hours",
    "LC 362 Design Hit Counter (M) • LC 1603 Design Parking System (E) • LC 295 Find Median from Data Stream (H)",
    "Easy-Hard", "3 timed design solutions", 3)

# ── MOCK INTERVIEWS & REVISION (Days 131-150) ──────────────────
MK = "Mock Interviews & Revision"
dsa(MK, "Blind 75 – Review Day 1 (Arrays/Hashing/Two Ptr)",
    "Review Blind 75 problems for Arrays, Hashing, Two Pointers",
    "Re-solve 6 Blind 75 problems from memory",
    "LC 1 Two Sum • LC 49 Group Anagrams • LC 128 Longest Consecutive • LC 238 Product of Array • LC 15 3Sum • LC 11 Container Water",
    "Med", "6 solutions from memory", 3)

dsa(MK, "Blind 75 – Review Day 2 (Sliding Window/Binary Search)",
    "Review Blind 75 problems for Sliding Window, Binary Search",
    "Re-solve 6 Blind 75 problems from memory",
    "LC 3 Longest Substring • LC 76 Min Window Substring • LC 424 Longest Repeating • LC 33 Search Rotated • LC 153 Find Min Rotated • LC 875 Koko Bananas",
    "Med-Hard", "6 solutions from memory", 3)

dsa(MK, "Blind 75 – Review Day 3 (Stacks/LL/Trees)",
    "Review Blind 75 problems for Stack, Linked List, Trees",
    "Re-solve 6 Blind 75 problems from memory",
    "LC 20 Valid Parens • LC 84 Largest Rectangle • LC 206 Reverse LL • LC 23 Merge K Lists • LC 226 Invert BT • LC 124 Max Path Sum",
    "Med-Hard", "6 solutions from memory", 3)

dsa(MK, "Blind 75 – Review Day 4 (Graphs/DP)",
    "Review Blind 75 problems for Graphs, Dynamic Programming",
    "Re-solve 6 Blind 75 problems from memory",
    "LC 200 Number of Islands • LC 133 Clone Graph • LC 207 Course Schedule • LC 70 Climbing Stairs • LC 322 Coin Change • LC 300 LIS",
    "Med-Hard", "6 solutions from memory", 3)

dsa(MK, "Blind 75 – Review Day 5 (DP/Greedy/Intervals)",
    "Review remaining Blind 75 problems",
    "Re-solve 6 Blind 75 problems from memory",
    "LC 152 Max Product Subarray • LC 416 Partition Equal Subset • LC 62 Unique Paths • LC 55 Jump Game • LC 56 Merge Intervals • LC 57 Insert Interval",
    "Med", "6 solutions from memory", 3)

dsa(MK, "Mock Interview 1 – Easy/Medium",
    "Simulate 45-min MAANG phone screen: 2 problems, think aloud",
    "Set timer; solve 2 unseen medium problems with verbal explanation",
    "Pick 2 random mediums from Neetcode 150",
    "Medium", "2 solutions + self-recorded think-aloud notes", 3)

dsa(MK, "Mock Interview 2 – Medium/Hard",
    "Simulate 45-min MAANG onsite round: 1 medium + 1 hard",
    "Set timer; solve 1 medium + 1 hard with optimization discussion",
    "Pick 1 medium + 1 hard from Neetcode 150",
    "Med-Hard", "2 solutions + complexity analysis + optimization notes", 3)

dsa(MK, "Mock Interview 3 – System Design Coding",
    "Simulate design round: implement a complex data structure",
    "Set timer; implement a full design problem in 45 minutes",
    "Pick 1 hard design problem (LRU/LFU/Twitter/File System)",
    "Hard", "1 complete design implementation", 3)

dsa(MK, "Weak Topic Review – Day 1",
    "Identify 3 weakest DSA topics from practice history",
    "Study theory + solve 4 problems from weakest topic",
    "4 problems from weakest area",
    "Med-Hard", "4 solutions + revised theory notes", 3)

dsa(MK, "Weak Topic Review – Day 2",
    "Continue weak topic drilling",
    "Solve 4 more problems from 2nd/3rd weakest areas",
    "4 problems from weak areas",
    "Med-Hard", "4 solutions + pattern notes", 3)

dsa(MK, "Mock Interview 4 – Full Loop Sim (Round 1)",
    "Simulate coding round 1: arrays/strings/hashing",
    "2 problems in 45 minutes, explain approach before coding",
    "2 random array/string/hash problems",
    "Medium", "2 solutions with interviewer-style notes", 3)

dsa(MK, "Mock Interview 5 – Full Loop Sim (Round 2)",
    "Simulate coding round 2: trees/graphs",
    "2 problems in 45 minutes, identify BFS vs DFS approach",
    "2 random tree/graph problems",
    "Med-Hard", "2 solutions with approach justification", 3)

dsa(MK, "Mock Interview 6 – Full Loop Sim (Round 3)",
    "Simulate coding round 3: DP",
    "2 problems in 60 minutes, explain state transitions clearly",
    "2 random DP problems",
    "Med-Hard", "2 solutions with state transition diagrams", 3)

dsa(MK, "Speed Round – Easy Blitz",
    "Speed solving: build confidence and pattern recognition",
    "Solve 10 easy problems in 90 minutes (≤9 min each)",
    "10 random easy problems from LeetCode",
    "Easy", "10 solutions + time log per problem", 3)

dsa(MK, "Speed Round – Medium Blitz",
    "Speed solving: medium difficulty pattern recognition",
    "Solve 6 medium problems in 2 hours (≤20 min each)",
    "6 random medium problems from LeetCode",
    "Medium", "6 solutions + time log per problem", 3)

dsa(MK, "Contest Practice – Day 1",
    "Simulate LeetCode weekly contest conditions",
    "Solve 4 contest problems in 90 minutes (1E + 2M + 1H)",
    "Use a past LeetCode weekly contest",
    "Mixed", "4 solutions + score analysis", 3)

dsa(MK, "Contest Practice – Day 2",
    "Simulate another LeetCode weekly contest",
    "Solve 4 contest problems in 90 minutes",
    "Use a different past LeetCode weekly contest",
    "Mixed", "4 solutions + improvement notes vs Day 1", 3)

dsa(MK, "Final Mock – Google Style",
    "Full 45-min Google-style interview simulation",
    "1-2 problems, optimal solution, test cases, edge cases, complexity",
    "Random problems matching Google tag on LeetCode",
    "Med-Hard", "Solutions + detailed interview debrief", 3)

dsa(MK, "Final Mock – Meta Style",
    "Full 45-min Meta-style interview simulation",
    "1-2 problems, clean code, handle follow-ups, BFS/graph emphasis",
    "Random problems matching Meta tag on LeetCode",
    "Med-Hard", "Solutions + detailed interview debrief", 3)

dsa(MK, "Final Review & Cheat Sheet",
    "Create master cheat sheet of all patterns, templates, and gotchas",
    "Compile all patterns, time complexities, and templates into one reference",
    "N/A — consolidation day",
    "Theory", "DSA master cheat sheet (all patterns + templates + complexities)", 3)


# ════════════════════════════════════════════════════════════════
# BUILD THE WORKBOOK
# ════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()

# ────── Sheet 1: Day-by-Day DSA Plan ───────────────────────────
ws = wb.active
ws.title = "DSA Day-by-Day Plan"

dsa_headers = [
    "Day #", "Week #", "Category", "Topic",
    "Learning Objectives", "Implementation / Coding Task",
    "LeetCode Problems", "Difficulty", "Deliverable",
    "Est. Hours", "Status", "Notes"
]

dsa_col_widths = [7, 8, 26, 38, 52, 52, 64, 11, 44, 10, 12, 30]

for col_idx, (hdr, w) in enumerate(zip(dsa_headers, dsa_col_widths), 1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w

status_dv = DataValidation(
    type="list",
    formula1='"Not Started,In Progress,Completed,Skipped"',
    allow_blank=True,
)
status_dv.error = "Pick a valid status"
status_dv.errorTitle = "Invalid Status"
ws.add_data_validation(status_dv)

for (day, cat, topic, learn, impl, problems, diff, deliv, hrs) in DSA_PLAN:
    row = day + 1
    week = (day - 1) // 7 + 1

    ws.cell(row=row, column=1, value=day).alignment = CENTER_ALIGN
    ws.cell(row=row, column=2, value=week).alignment = CENTER_ALIGN
    ws.cell(row=row, column=3, value=cat)
    ws.cell(row=row, column=4, value=topic)
    ws.cell(row=row, column=5, value=learn)
    ws.cell(row=row, column=6, value=impl)
    ws.cell(row=row, column=7, value=problems)
    ws.cell(row=row, column=8, value=diff).alignment = CENTER_ALIGN
    ws.cell(row=row, column=9, value=deliv)
    ws.cell(row=row, column=10, value=hrs).alignment = CENTER_ALIGN
    sc = ws.cell(row=row, column=11, value="Not Started")
    sc.alignment = CENTER_ALIGN
    status_dv.add(sc)
    ws.cell(row=row, column=12, value="")

    cat_color = DSA_TOPIC_COLORS.get(cat, "2C3E50")
    cat_fill = PatternFill("solid", fgColor=cat_color)

    for col in range(1, 13):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
        c.font = BODY_FONT
        if col in (5, 6, 7, 9):
            c.alignment = WRAP_ALIGN

    ws.cell(row=row, column=3).fill = cat_fill
    ws.cell(row=row, column=3).font = PHASE_FONT

ws.auto_filter.ref = f"A1:L{len(DSA_PLAN)+1}"
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = "8E44AD"

# ────── Sheet 2: Topic Summary ─────────────────────────────────
ws2 = wb.create_sheet("Topic Summary")

sum_headers = ["Category", "Days", "Total Hours", "# LeetCode Problems (approx)", "Key Patterns"]
sum_widths = [28, 10, 12, 26, 75]

for col_idx, (hdr, w) in enumerate(zip(sum_headers, sum_widths), 1):
    cell = ws2.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    ws2.column_dimensions[get_column_letter(col_idx)].width = w

dsa_focus = {
    "Foundations": "Big-O, recursion, hash table internals, Python DS mastery, problem-solving framework",
    "Arrays & Hashing": "Prefix sum, frequency counting, grouping, two sum variants, matrix traversal",
    "Two Pointers": "Opposite ends, same direction, partition, Dutch National Flag, multi-pointer",
    "Sliding Window": "Fixed window, variable window, window + hashmap, min/max window problems",
    "Binary Search": "3 templates, sorted array variants, search on answer space, 2D search",
    "Stacks & Queues": "Monotonic stack, next greater/smaller, parentheses, decode, calculator",
    "Linked Lists": "Fast/slow pointer, reversal patterns, merge/sort, LRU/LFU cache",
    "Sorting": "Merge/Quick/Heap sort, counting/radix sort, custom sort, interval sorting",
    "Trees": "DFS/BFS traversals, path problems, BST operations, construction, serialization",
    "Tries": "Trie implementation, word search, auto-complete, prefix matching",
    "Heaps / Priority Queues": "Top-K, two heaps (median), merge K sorted, scheduling",
    "Backtracking": "Subsets, permutations, combinations, grid/board, string partition, N-Queens",
    "Graphs": "BFS, DFS, topological sort, Union-Find, Dijkstra, Bellman-Ford, MST, bipartite",
    "Dynamic Programming": "1D, 2D grid, string, knapsack, stocks, interval, bitmask, tree, digit DP",
    "Greedy": "Greedy choice property, intervals, scheduling, string, Kadane's",
    "Intervals": "Merge, insert, sweep line, meeting rooms, job scheduling",
    "Bit Manipulation": "XOR patterns, bitmask, set/clear/toggle bits, missing/duplicate",
    "Math & Geometry": "Number theory, primes (sieve), combinatorics, coordinate geometry",
    "Design": "LRU/LFU cache, randomized set, time-based KV, iterator, complex OO design",
    "Mock Interviews & Revision": "Blind 75 review, mock interviews, speed rounds, contest practice, cheat sheets",
}

cat_data: dict[str, dict] = {}
for (day, cat, topic, learn, impl, problems, diff, deliv, hrs) in DSA_PLAN:
    if cat not in cat_data:
        cat_data[cat] = {"count": 0, "hours": 0.0, "problems": 0}
    cat_data[cat]["count"] += 1
    cat_data[cat]["hours"] += hrs
    bullet_count = problems.count("•") + (1 if problems.strip() and "N/A" not in problems else 0)
    cat_data[cat]["problems"] += bullet_count

for row_idx, (cat, data) in enumerate(cat_data.items(), start=2):
    ws2.cell(row=row_idx, column=1, value=cat)
    ws2.cell(row=row_idx, column=2, value=data["count"]).alignment = CENTER_ALIGN
    ws2.cell(row=row_idx, column=3, value=data["hours"]).alignment = CENTER_ALIGN
    ws2.cell(row=row_idx, column=4, value=data["problems"]).alignment = CENTER_ALIGN
    ws2.cell(row=row_idx, column=5, value=dsa_focus.get(cat, ""))

    cat_color = DSA_TOPIC_COLORS.get(cat, "2C3E50")
    ws2.cell(row=row_idx, column=1).fill = PatternFill("solid", fgColor=cat_color)
    ws2.cell(row=row_idx, column=1).font = PHASE_FONT
    for col in range(1, 6):
        ws2.cell(row=row_idx, column=col).border = THIN_BORDER
        ws2.cell(row=row_idx, column=col).font = BODY_FONT
        if col == 5:
            ws2.cell(row=row_idx, column=col).alignment = WRAP_ALIGN

# Totals row
total_r = len(cat_data) + 2
ws2.cell(row=total_r, column=1, value="TOTAL").font = Font(bold=True, size=11, name="Calibri")
ws2.cell(row=total_r, column=2, value=sum(d["count"] for d in cat_data.values())).font = Font(bold=True, size=11, name="Calibri")
ws2.cell(row=total_r, column=3, value=sum(d["hours"] for d in cat_data.values())).font = Font(bold=True, size=11, name="Calibri")
ws2.cell(row=total_r, column=4, value=sum(d["problems"] for d in cat_data.values())).font = Font(bold=True, size=11, name="Calibri")
for col in range(1, 6):
    ws2.cell(row=total_r, column=col).border = THIN_BORDER
    ws2.cell(row=total_r, column=col).alignment = CENTER_ALIGN

ws2.freeze_panes = "A2"
ws2.sheet_properties.tabColor = "9B59B6"

# ────── Sheet 3: Pattern Checklist ─────────────────────────────
ws3 = wb.create_sheet("Pattern Checklist")

pat_headers = ["#", "Pattern", "Key Template / Technique", "Must-Know Problems", "Confident?", "Notes"]
pat_widths = [5, 28, 50, 60, 12, 30]

for col_idx, (hdr, w) in enumerate(zip(pat_headers, pat_widths), 1):
    cell = ws3.cell(row=1, column=col_idx, value=hdr)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER_ALIGN
    cell.border = THIN_BORDER
    ws3.column_dimensions[get_column_letter(col_idx)].width = w

patterns = [
    ("Two Pointers – Opposite", "while left < right: adjust based on sum/condition", "LC 15 3Sum • LC 11 Container Water • LC 42 Trapping Rain Water"),
    ("Two Pointers – Same Dir", "Fast/slow pointer for cycle, middle, nth from end", "LC 141 Cycle • LC 876 Middle • LC 19 Remove Nth"),
    ("Sliding Window – Fixed", "Maintain window of size K; slide by adding right, removing left", "LC 643 Max Avg Subarray • LC 1456 Max Vowels"),
    ("Sliding Window – Variable", "Expand right, shrink left while invalid; track best", "LC 3 Longest Substring • LC 76 Min Window Substring"),
    ("Binary Search – Classic", "left, right = 0, n-1; mid = (l+r)//2; adjust boundaries", "LC 704 Binary Search • LC 33 Search Rotated"),
    ("Binary Search – On Answer", "Binary search on answer space with feasibility check", "LC 875 Koko Bananas • LC 1011 Ship Packages"),
    ("Monotonic Stack", "Stack maintains increasing/decreasing order; pop when condition met", "LC 739 Daily Temps • LC 84 Largest Rectangle"),
    ("BFS – Level Order", "Queue-based; process level by level; shortest path in unweighted", "LC 994 Rotting Oranges • LC 127 Word Ladder"),
    ("DFS – Recursive", "Visit node, mark visited, recurse on neighbors", "LC 200 Islands • LC 695 Max Area"),
    ("DFS – Iterative", "Stack-based DFS for explicit control", "LC 200 Islands variant"),
    ("Topological Sort (Kahn's)", "In-degree map + queue; process nodes with 0 in-degree", "LC 207 Course Schedule • LC 210 Course Schedule II"),
    ("Union-Find", "find() with path compression + union() by rank", "LC 323 Components • LC 684 Redundant Connection"),
    ("Dijkstra", "Min-heap + dist array; relax edges greedily", "LC 743 Network Delay • LC 787 Cheapest Flights"),
    ("Backtracking – Subsets", "for i in range(start, n): choose → explore → unchoose", "LC 78 Subsets • LC 90 Subsets II"),
    ("Backtracking – Permutations", "for i in range(n): if not used[i]: choose → explore → unchoose", "LC 46 Permutations • LC 47 Permutations II"),
    ("Backtracking – Combinations", "Combination sum with pruning (sort + skip duplicates)", "LC 39 Combination Sum • LC 40 Combination Sum II"),
    ("DP – 1D Bottom-Up", "dp[i] = f(dp[i-1], dp[i-2], ...); iterate left to right", "LC 70 Climbing Stairs • LC 198 House Robber"),
    ("DP – 2D Grid", "dp[i][j] = f(dp[i-1][j], dp[i][j-1]); process row by row", "LC 62 Unique Paths • LC 64 Min Path Sum"),
    ("DP – String (LCS)", "dp[i][j] = LCS of s1[:i] and s2[:j]", "LC 1143 LCS • LC 72 Edit Distance"),
    ("DP – Knapsack (0/1)", "dp[i][w] = max(dp[i-1][w], dp[i-1][w-wt[i]] + val[i])", "LC 416 Partition Equal Subset • LC 494 Target Sum"),
    ("DP – Stock State Machine", "States: hold, sold, cooldown; transition between states", "LC 121/122/123/188/309 Stock variants"),
    ("DP – Interval", "dp[i][j] = optimal for subarray i..j; try all split points k", "LC 312 Burst Balloons • LC 1039 Triangulation"),
    ("Trie", "TrieNode with children dict + is_end; insert/search/startsWith", "LC 208 Implement Trie • LC 212 Word Search II"),
    ("Heap – Top K", "Min-heap of size K; push and pop to maintain top K", "LC 215 Kth Largest • LC 347 Top K Frequent"),
    ("Heap – Two Heaps (Median)", "Max-heap for lower half + min-heap for upper half; balance sizes", "LC 295 Find Median • LC 480 Sliding Window Median"),
    ("Greedy – Interval Scheduling", "Sort by end time; greedily select non-overlapping", "LC 435 Non-overlapping • LC 452 Min Arrows"),
    ("Merge Intervals", "Sort by start; merge if overlap; else append", "LC 56 Merge Intervals • LC 57 Insert Interval"),
    ("Bit Manipulation – XOR", "a^a=0, a^0=a; find missing/single number", "LC 136 Single Number • LC 268 Missing Number"),
]

yn_dv = DataValidation(type="list", formula1='"Yes,No,Partial"', allow_blank=True)
ws3.add_data_validation(yn_dv)

for i, (pattern, template, must_know) in enumerate(patterns, start=1):
    row = i + 1
    ws3.cell(row=row, column=1, value=i).alignment = CENTER_ALIGN
    ws3.cell(row=row, column=2, value=pattern)
    ws3.cell(row=row, column=3, value=template)
    ws3.cell(row=row, column=4, value=must_know)
    c = ws3.cell(row=row, column=5, value="No")
    c.alignment = CENTER_ALIGN
    yn_dv.add(c)
    ws3.cell(row=row, column=6, value="")
    for col in range(1, 7):
        ws3.cell(row=row, column=col).border = THIN_BORDER
        ws3.cell(row=row, column=col).font = BODY_FONT
        if col in (3, 4):
            ws3.cell(row=row, column=col).alignment = WRAP_ALIGN

ws3.auto_filter.ref = f"A1:F{len(patterns)+1}"
ws3.freeze_panes = "A2"
ws3.sheet_properties.tabColor = "E74C3C"

# ────── Save ────────────────────────────────────────────────────
output_path = "/home/pal/Projects/MachineLearning/DSA_Interview_Prep_Plan.xlsx"
wb.save(output_path)
print(f"✅ Workbook saved to: {output_path}")
print(f"   Total learning days: {len(DSA_PLAN)}")
print(f"   Total hours:         {sum(r[8] for r in DSA_PLAN)}")
print(f"   LeetCode problems:   ~{sum(d['problems'] for d in cat_data.values())}")
print(f"   Sheets: {wb.sheetnames}")
